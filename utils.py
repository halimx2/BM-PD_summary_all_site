from options   import KIND_OPTIONS, SITE_OPTIONS, PROCESS_OPTIONS, UNIT_OPTIONS
from load_data import load_sheet_data


import io
import pandas as pd
import streamlit as st
from datetime import datetime

__all__ = [
    "load_sheet_data",
    "KIND_OPTIONS",
    "SITE_OPTIONS",
    "PROCESS_OPTIONS",
    "UNIT_OPTIONS",
    "transform_to_WA_schema", "to_excel_template_WA"
]


def _to_dt(s):
    return pd.to_datetime(s, errors="coerce")

def _fmt_hhmm(dt):
    if pd.isna(dt):
        return ""
    # OS 독립적으로 "H:MM" 포맷
    return f"{int(dt.hour)}:{int(dt.minute):02d}"

def _elapsed_hhmm(start, end):
    if pd.isna(start) or pd.isna(end):
        return ""
    mins = int((end - start).total_seconds() // 60)
    h, m = divmod(mins, 60)
    return f"{h}:{m:02d}"


def _proc_label(x):
    # 공정 라벨(선택) — 화면처럼 앞에 ①,② 붙이고 싶을 때 사용
    PROCESS_LABEL = {
        "Inspection": "①Inspection",
        "Lamination": "①Lamination",
        "Electrode Supply":"①Lamination",
        "Stacking":   "②Stacking",
        "D-Stacking":   "②Stacking",
        "Taping": "②Stacking",
        "Taper": "②Stacking",
        "Inspection": "②Stacking",
    }
    if pd.isna(x): 
        return ""
    return PROCESS_LABEL.get(str(x), str(x))  # 없으면 원문 사용


def _bmpd_label(x):
    # BM/PD 라벨(선택)
    BMPD_LABEL = {"BM": "①BM", "PD": "②PD"}

    if pd.isna(x):
        return ""
    s = str(x).strip().upper()
    return BMPD_LABEL.get(s, s)

def transform_to_WA_schema(df_src: pd.DataFrame) -> pd.DataFrame:
    """
    입력 컬럼 예:
        종류, Site, 호기, Machine, Unit, Assy', 발생시간, 조치완료,
        조치 진행 시간(분), 작업자, 현상, 원인, 조치, 발생일
    """
    df = df_src.copy()

    dt_occ  = _to_dt(df.get("발생시간"))
    dt_done = _to_dt(df.get("조치완료"))
    dt_day  = _to_dt(df.get("발생일"))

    # 파생
    등록일 = dt_day.dt.date.astype("string").fillna("")
    # 주차   = dt_day.dt.isocalendar().week.astype("Int64") + 'W'
    # 주차   = dt_day.dt.isocalendar().week.astype("Int64") + 'W'
    주차 = dt_day.dt.isocalendar().week.map(lambda w: f"{w}W")
    월일   = dt_day.dt.strftime("%Y-%m-%d").fillna("")

    조치_start = [_fmt_hhmm(x) for x in dt_occ]
    조치_end   = [_fmt_hhmm(x) for x in dt_done]
    소요시간    = [_elapsed_hhmm(s, e) for s, e in zip(dt_occ, dt_done)]

    # 색지도 기준(샘플처럼 STOP=Start, START=End, 부동시간=소요시간)
    STOP   = 조치_start
    START  = 조치_end
    부동    = 소요시간

    공정      = df.get("Machine", "").map(_proc_label).astype("string")
    설비호기   = '#'+ df.get("호기", "").astype("string")
    bm_pd    = df.get("종류", "").map(_bmpd_label).astype("string")

    현상      = df.get("현상", "").astype("string")
    조치내용   = df.get("조치", "").astype("string")
    원인      = df.get("원인", "").astype("string")
    협력사    = df.get("협력사", "")
    조치자    = df.get("작업자", "")  # 원본 '작업자'를 조치자로 사용

    out = pd.DataFrame({
        "등록일": 등록일,
        "주차": 주차,
        "월/일": 월일,
        "조치시간 \n(Start)": 조치_start,
        "조치시간 \n(End)":   조치_end,
        "소요시간":           소요시간,
        "STOP\n(색지도 기준)":  STOP,
        "START\n(색지도 기준)": START,
        "부동시간\n(색지도 기준)": 부동,
        "공정":               공정,
        "설비\n호기":          설비호기,
        "BM/PD":             bm_pd,
        "BM/PD\n분류":        None,         # 현재 원본에 없음 → 공란
        "BM/PD\n세부항목":     None,         # 현재 원본에 없음 → 공란
        "대분류":             None,         # 현재 원본에 없음 → 공란
        "중분류":             None,         # 현재 원본에 없음 → 공란
        "소분류":             None,         # 현재 원본에 없음 → 공란
        "현상 Phenomenon\n(제품 모습, 설비 모습)": 현상,
        "조치 내용\n(교체 부품 반드시 기재)":       조치내용,
        "원인 또는 예상 원인\n(필요시 개선점 기재)":   원인,
        "협력사":            협력사,
        "조치자":            조치자,
    })

    # 최종 컬럼 순서(질문 이미지 순서 그대로)
    columns_out = [
        "등록일","주차","월/일",
        "조치시간 \n(Start)","조치시간 \n(End)","소요시간",
        "STOP\n(색지도 기준)","START\n(색지도 기준)","부동시간\n(색지도 기준)",
        "공정","설비\n호기","BM/PD","BM/PD\n분류","BM/PD\n세부항목",
        "대분류","중분류","소분류",
        "현상 Phenomenon\n(제품 모습, 설비 모습)",
        "조치 내용\n(교체 부품 반드시 기재)",
        "원인 또는 예상 원인\n(필요시 개선점 기재)",
        "협력사","조치자"
    ]

    return out[columns_out]



def to_excel_template_WA(df_final: pd.DataFrame) -> bytes:
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    import numpy as np
    buf = io.BytesIO()

    # 🔧 1. 엑셀 템플릿 로드
    wb = load_workbook("template/bmpd_daily_issue_WA.xlsx")
    ws = wb[wb.sheetnames[0]]

    # 🔧 2. 모든 결측값을 None으로 바꿔줌 (<NA>, NaT, np.nan 모두)
    clean = df_final.copy(deep=True)
    clean = clean.replace({pd.NA: None, np.nan: None})
    clean = clean.astype(object).where(pd.notna(clean), None)

    # 🔧 3. 기존 내용 삭제 후 데이터 다시 쓰기
    ws.delete_rows(1, ws.max_row)

    # dataframe_to_rows를 쓰면 자동으로 각 셀에 맞는 타입으로 넣어줌
    for r in dataframe_to_rows(clean, index=False, header=True):
        ws.append(r)

    # 🔧 4. 메모리로 저장
    wb.save(buf)
    return buf.getvalue()

