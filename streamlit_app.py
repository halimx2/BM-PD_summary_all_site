import streamlit as st
import pandas as pd
import io
import os
from openpyxl import load_workbook
from report_extractor import parse_chat_text, extract_report_data

# --- 기본 설정 ---
st.set_page_config(page_title="BM, PD 이력 정리", layout="wide")
st.title("📋 BM, PD 이력 정리")

# --- 파일 업로드 ---
st.subheader("📂 채팅 리포트 파일 업로드")
txt_files = st.file_uploader("채팅 txt 파일 업로드 (여러 개 가능)", type="txt", accept_multiple_files=True)

if txt_files:
    all_messages = []
    for f in txt_files:
        try:
            text = f.read().decode('utf-8')
        except Exception:
            text = f.read().decode('cp949', errors='ignore')
        all_messages.extend(parse_chat_text(text))
    df_raw = extract_report_data(all_messages)

    if df_raw.empty:
        st.warning("유효한 리포트 데이터가 없습니다.")
    else:
        st.success(f"Raw Data {len(df_raw)}건 추출 완료")
        st.dataframe(df_raw)

        TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'template.xlsx')
        try:
            wb = load_workbook(TEMPLATE_PATH)
        except Exception as e:
            st.error(f"템플릿 로드 오류: {e}")
        else:
            sheet_name = 'RawData'
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name, 0)
            if ws.max_row > 1:
                ws.delete_rows(idx=2, amount=ws.max_row - 1)
            for r, row in enumerate(df_raw.itertuples(index=False, name=None), start=2):
                for c, val in enumerate(row, start=1):
                    ws.cell(row=r, column=c, value=val)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            st.download_button(
                label="엑셀 다운로드",
                data=output.getvalue(),
                file_name=f"analysis_{pd.Timestamp.now().strftime('%y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
else:
    st.info("채팅 txt 파일을 업로드해주세요.")
